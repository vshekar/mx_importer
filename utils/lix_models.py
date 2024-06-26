import re
from collections import defaultdict
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from PIL import Image as PILImage
from qrcode.constants import ERROR_CORRECT_L
from qrcode.main import QRCode
from qtpy.QtCore import Qt, Signal
from qtpy.QtGui import QColor

from utils.pandas_model import BasePandasModel


class LIXPlatePandasModel(BasePandasModel):
    def __init__(self, dataframe: pd.DataFrame, parent=None, well_name=None) -> None:
        self.non_editable_cells = set()
        self.well_name = well_name
        super().__init__(dataframe, parent)

        self.dataChanged.connect(self.updateDelegateItems)

    def flags(self, index):
        defaultFlags = super(BasePandasModel, self).flags(index)
        if (index.row(), index.column()) not in self.non_editable_cells:
            return (
                Qt.ItemFlag.ItemIsSelectable
                | Qt.ItemFlag.ItemIsEnabled
                | Qt.ItemFlag.ItemIsEditable
            )
        else:
            return defaultFlags & ~Qt.ItemFlag.ItemIsEditable

    def preprocessData(self):
        if "Stock" in self._dataframe.columns:
            self._dataframe = self._dataframe[~self._dataframe["Well"].isnull()][
                ["Well", "Sample", "Buffer", "Volume (uL)", "Mixing", "Stock", "Notes"]
            ]

    def validateData(self, config):
        self.resetColors()
        self.checkSampleDuplicates(self._dataframe)
        # self.checkAll(self._dataframe)

    def check_sample_name(self, sample_name):
        if len(sample_name) > 42:  # file name length limit for Pilatus detectors
            print("Error: the sample name is too long:", len(sample_name))
            return False
        l1 = re.findall("[^:._A-Za-z0-9\-]", sample_name)
        if len(l1) > 0:
            print("Error: the file name contain invalid characters: ", l1)
            return False

        return True

    def checkAll(self, df1: pd.DataFrame, b_lim=3, v_margin=5):
        # List of checks
        # Sample name limits <= 42
        # Invalid characters re match "[^:._A-Za-z0-9\-]"
        # Repeating sample names

        # check redundant sample name, and sample name validity
        all_samples = list(df1["Sample"])
        all_buffers = list(df1[~df1["Buffer"].isnull()]["Buffer"])
        all_sample_wells = list(df1["Well"])

        for sample_name, count in df1["Sample"].value_counts().items():
            if not self.check_sample_name(sample_name):
                self._changeCellColors(
                    df1.columns.get_loc("Sample"),
                    df1[df1["Sample"] == sample_name].index,
                )
                raise TypeError(f"Invalid sample names, highlighted in red")
            if count > 1:
                self._changeCellColors(
                    df1.columns.get_loc("Sample"),
                    df1[df1["Sample"] == sample_name].index,
                )
                raise TypeError(f"Redundant sample names, highlighted in red")

        # check buffer name, and how many samples use the same buffer for subtraction
        for buffer_name, count in (
            df1[~df1["Buffer"].isnull()]["Buffer"].value_counts().items()
        ):
            if buffer_name not in df1["Sample"].values:
                self._changeCellColors(
                    df1.columns.get_loc("Buffer"),
                    df1[df1["Buffer"] == buffer_name].index,
                )
                raise TypeError(
                    f"'{buffer_name}' is not a sample and therefore not a valid buffer."
                )
            if count > b_lim:
                self._changeCellColors(
                    df1.columns.get_loc("Buffer"),
                    df1[df1["Buffer"] == buffer_name].index,
                )
                raise TypeError(
                    f"'{buffer_name}' is used more than {b_lim} times for buffer subtraction."
                )

        sdict = df1.set_index("Sample").T.to_dict()
        # check whether every sample has a buffer and in the same row
        for sn in list(set(all_samples) - set(all_buffers)):
            if not isinstance(sdict[sn]["Buffer"], str):  # nan
                print(f"Warning: {sn} does not have a corresponding sample/buffer.")
            else:
                swell = sdict[sn]["Well"]
                bwell = sdict[sdict[sn]["Buffer"]]["Well"]
                if swell[0] != bwell[0]:
                    raise TypeError(
                        f"sample and buffer not in the same row: {swell} vs {bwell}"
                    )

        wdict = df1[~df1["Sample"].isnull()].set_index("Well").T.to_dict()
        wlist = list(wdict.keys())
        # each well either need to have volume specified, or how to generate the sample
        mlist = {}  # source well for each new sample
        slist = {}  # total volume out of source well
        for wn in wlist:
            if isinstance(wdict[wn]["Mixing"], str):
                tl = {}
                for tt in wdict[wn]["Mixing"].strip().split(","):
                    w, v = tt.strip().split(":")
                    if w not in wlist:
                        raise TypeError(f"source well {w} is empty.")
                    if not isinstance(wdict[w]["Stock"], str):
                        raise TypeError(
                            f"source well {w} is not designated as a stock well."
                        )
                    if w in tl.keys():
                        raise TypeError(
                            f"{w} appears more than once in the mixing list for {wn}."
                        )
                    tl[w] = float(v)
                    if not w in slist.keys():
                        slist[w] = float(v)
                    else:
                        slist[w] += float(v)
                mlist[wn] = tl
            elif np.isnan(wdict[wn]["Volume (uL)"]):
                raise TypeError(f"neither volume nor mixing is specified for well {wn}")

        for wn in slist.keys():
            if slist[wn] + v_margin > wdict[wn]["Volume (uL)"]:
                raise TypeError(
                    f"not sufficient sample in well {wn}, need {slist[wn]+v_margin}"
                )

        rdict = {}
        for w in all_sample_wells:
            rn = w[0]
            if rn in rdict.keys():
                rdict[rn] += [w]
            else:
                rdict[rn] = [w]

        rlen = np.asarray([len(rdict[rn]) for rn in rdict.keys()])
        if len(rlen[rlen < 5]) > 0:
            print(f"Consider consolidating the rows, more than two rows are half full.")

    def checkValidTemplate(self):
        if len(self._dataframe) != 12:
            raise TypeError(
                f"Invalid number of rows in table, expected 12, found {len(self._dataframe)}"
            )
        if len(self._dataframe.columns) != 7:
            raise TypeError(
                f"Invalid number of columns in table, expected 7, found {len(self._dataframe.columns)}"
            )
        if (
            len(
                column_diff := set(
                    [
                        "Well",
                        "Sample",
                        "Buffer",
                        "Volume (uL)",
                        "Mixing",
                        "Stock",
                        "Notes",
                    ]
                ).difference(set(self._dataframe.columns))
            )
            != 0
        ):
            raise TypeError(f"Missing columns : {column_diff}")

    def checkSampleDuplicates(self, dataframe: pd.DataFrame):
        duplicates = dataframe[~pd.isna(dataframe["Sample"])]["Sample"].duplicated(
            keep=False
        )
        duplicate_indexes = dataframe.index[
            (dataframe["Sample"] != "") & duplicates
        ].to_numpy()
        # Next few lines are to fix the mismatch between the index of the dataframe (based on the excel file)
        # Whereas index in each model starts from 0
        offset = ord(self.well_name) - ord("A")
        duplicate_indexes = duplicate_indexes - (offset * 12)
        column_index = dataframe.columns.get_loc("Sample")
        if len(duplicate_indexes):
            self._changeCellColors(
                column_index, duplicate_indexes, color=QColor(Qt.GlobalColor.red)
            )
            return False
        return True

    def _changeCellColors(
        self, column_index: int, row_indices, color=QColor(Qt.GlobalColor.red)
    ) -> None:
        for idx in row_indices:
            self.changeColor(idx, column_index, color)

    def updateDelegateItems(self, topLeft, bottomRight, roles):
        changedColumn = topLeft.column()
        changedRow = topLeft.row()

        if changedColumn == self.getColumnIndex("Stock"):
            mixing_column = self.getColumnIndex("Mixing")
            if self._dataframe.iat[changedRow, changedColumn] == "X":
                self.non_editable_cells.add((changedRow, mixing_column))
            else:
                self.non_editable_cells.discard((changedRow, mixing_column))

        if changedColumn == self.getColumnIndex(
            "Volume (uL)"
        ) or changedColumn == self.getColumnIndex("Mixing"):
            wells = self._dataframe[~self._dataframe["Stock"].isnull()][
                "Well"
            ].to_list()
            total_volumes = self._dataframe[~self._dataframe["Stock"].isnull()][
                "Volume (uL)"
            ].to_list()
            total_volumes = {well: vol for well, vol in zip(wells, total_volumes)}

    def getColumnIndex(self, column_name):
        return self._dataframe.columns.get_loc(column_name)

    def get_mixing_data(self, index):
        used_volumes = defaultdict(int)
        for i, text in enumerate(self._dataframe["Mixing"]):
            if not isinstance(text, str) or i == index.row():
                continue
            for token in text.split(","):
                if len(token.split(":")) == 2:
                    row, amount = token.strip().split(":")
                    used_volumes[row] += int(amount)

        total_volume = {
            row["Well"]: row["Volume (uL)"]
            for i, row in self._dataframe[self._dataframe["Stock"] == "X"].iterrows()
            if not pd.isna(row["Volume (uL)"])
        }

        remaining_volume = {
            well: int(total_volume[well]) - int(used_volumes[well])
            for well in total_volume.keys()
        }
        return remaining_volume

    def get_current_samples(self):
        return self._dataframe[
            self._dataframe["Stock"].isna() & ~self._dataframe["Sample"].isna()
        ]["Sample"].to_list()


class LIXHolderPandasModel(BasePandasModel):
    updated_holder_name = Signal(int, str)

    def __init__(
        self, dataframe: pd.DataFrame, holder_index, parent=None, holder_name=None
    ) -> None:
        self.holder_name = holder_name
        self.holder_index = holder_index
        self.editable_cells = {(0, 0)}

        super().__init__(dataframe, parent)
        self.dataChanged.connect(self.updateItems)

    def flags(self, index):
        defaultFlags = super(BasePandasModel, self).flags(index)
        if ((index.row(), index.column()) in self.editable_cells) or index.column() in (
            2,
            3,
            4,
        ):
            return (
                Qt.ItemFlag.ItemIsSelectable
                | Qt.ItemFlag.ItemIsEnabled
                | Qt.ItemFlag.ItemIsEditable
            )
        else:
            return defaultFlags & ~Qt.ItemFlag.ItemIsEditable

    def getColumnIndex(self, column_name):
        return self._dataframe.columns.get_loc(column_name)

    def updateItems(self, topLeft, bottomRight, roles):
        changedColumn = topLeft.column()
        changedRow = topLeft.row()

        if changedColumn == self.getColumnIndex("holderName") and changedRow == 0:
            self.holder_name = self._dataframe.iloc[changedRow, changedColumn]
            self.updated_holder_name.emit(self.holder_index, self.holder_name)

    def get_current_samples(self):
        return self._dataframe[~self._dataframe["sampleName"].isna()][
            "sampleName"
        ].unique()

    def setData(self, index, value, role: int) -> bool:
        if role == Qt.ItemDataRole.EditRole:
            column_type = self._dataframe.dtypes.iloc[index.column()]
            try:
                # Try to cast the value to the column type
                value = column_type.type(value)
            except (ValueError, TypeError):
                # If casting fails, do not update the DataFrame
                return False

            if index.column() == self.getColumnIndex("volume"):
                if self._dataframe["sampleName"].iloc[index.row()] == "":
                    return False

            self._dataframe.iloc[index.row(), index.column()] = value
            self.dataChanged.emit(index, index, [role])
            return True
        return False


def make_plate_QR_code(proposal_id, SAF_id, plate_id, path):
    """depends on blabel
    generate a pdf file, with plate outline
    """
    code = [str(proposal_id), str(SAF_id), str(plate_id)]

    str_in = "-".join(code)
    qr = QRCode(
        version=1,
        error_correction=ERROR_CORRECT_L,
        box_size=10,
        border=4,
    )
    qr.add_data(str_in)
    qr.make(fit=True)

    qr_img = qr.make_image(fill="black", back_color="white")
    qr_filepath = Path(path) / Path("qrcode.png")
    qr_img.save(qr_filepath)

    return str_in, qr_filepath


def write_excel(df, filepath, sheet_name, qr_filename):
    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)

    wb = load_workbook(filepath)
    ws = wb.create_sheet(title=f"{sheet_name} QR Code")

    """
    qr_pil_img = PILImage.open(qr_filename)
    qr_pil_img = qr_pil_img.resize(
        (int(qr_pil_img.width / 2), int(qr_pil_img.height / 2)), PILImage.ANTIALIAS
    )
    qr_pil_img.save(qr_filename)
    """

    # Insert the image
    img = Image(qr_filename)
    ws.add_image(img, "A1")

    # Save the workbook
    wb.save(filepath)
